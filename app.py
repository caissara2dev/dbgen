import csv
import io
import os
import zipfile
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile, status
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from openpyxl import Workbook

# Reuse helpers from existing scripts
from faker import Faker
from scripts.gerar_clientes import (
    gerar_celular,
    gerar_cpf,
    gerar_email_local,
    gerar_nome,
    gerar_codigo,
)
from scripts.gerar_compras import gerar_datas, gerar_filiais, montar_compras


# Configurables
MAX_CLIENTS = int(os.environ.get("DBGEN_MAX_CLIENTS", "50000"))
MAX_PURCHASES = int(os.environ.get("DBGEN_MAX_PURCHASES", "200000"))
MAX_UPLOAD_MB = int(os.environ.get("DBGEN_MAX_UPLOAD_MB", "5"))
ROOT_PATH = os.environ.get("ROOT_PATH", "/dbgen")  # to serve behind caissara.com/dbgen

security = HTTPBasic()


def require_basic(credentials: HTTPBasicCredentials = Depends(security)) -> None:
    """Simple HTTP Basic auth.

    - If DBGEN_PASSWORD is set, require it. If DBGEN_USER is set, require both.
    - If neither is set, auth is disabled (development mode).
    """
    cfg_user = os.environ.get("DBGEN_USER")
    cfg_pass = os.environ.get("DBGEN_PASSWORD")
    if not cfg_user and not cfg_pass:
        # Auth disabled
        return
    # Compare
    if cfg_user and credentials.username != cfg_user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Unauthorized",
            headers={"WWW-Authenticate": "Basic"},
        )
    if cfg_pass and credentials.password != cfg_pass:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Unauthorized",
            headers={"WWW-Authenticate": "Basic"},
        )


app = FastAPI(title="DB Generators", version="0.1.0", root_path=ROOT_PATH)


def _in_memory_clients(count: int, domain: str, seed: Optional[int]) -> tuple[bytes, bytes]:
    """Generate clients CSV and XLSX in memory.

    Returns (csv_bytes, xlsx_bytes)
    """
    if count < 1:
        raise HTTPException(status_code=400, detail="count must be >= 1")
    if count > MAX_CLIENTS:
        raise HTTPException(status_code=400, detail=f"count exceeds limit ({MAX_CLIENTS})")

    fake = Faker("pt_BR")
    if seed is not None:
        try:
            fake.seed_instance(seed)
        except Exception:
            pass

    usados_cpf: set[str] = set()
    usados_email: set[str] = set()
    header = ["CódigoCliente", "NomeCompleto", "Celular", "CPF", "Email"]
    rows: list[list[str]] = []

    for i in range(1, count + 1):
        nome = gerar_nome(fake)

        cpf = gerar_cpf()
        while cpf in usados_cpf:
            cpf = gerar_cpf()
        usados_cpf.add(cpf)

        local = gerar_email_local(nome)
        email = f"{local}@{domain}"
        n = 2
        while email in usados_email:
            email = f"{local}{n}@{domain}"
            n += 1
        usados_email.add(email)

        celular = gerar_celular()
        codigo = gerar_codigo(i)
        rows.append([codigo, nome, celular, cpf, email])

    # CSV
    csv_buf = io.StringIO(newline="")
    w = csv.writer(csv_buf)
    w.writerow(header)
    w.writerows(rows)
    csv_bytes = csv_buf.getvalue().encode("utf-8")

    # XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(header)
    for r in rows:
        ws.append(r)
    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_bytes = xlsx_buf.getvalue()
    return csv_bytes, xlsx_bytes


def _parse_client_codes_from_upload(file_bytes: bytes) -> list[str]:
    """Parse 'CódigoCliente' (or variant) from uploaded CSV bytes."""
    try:
        s = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        # Fallback (rare): latin-1
        s = file_bytes.decode("latin-1")
    reader = csv.DictReader(io.StringIO(s))
    field = None
    for cand in ("CódigoCliente", "CodigoCliente", "Codigo", "Código"):
        if reader.fieldnames and cand in reader.fieldnames:
            field = cand
            break
    if not field:
        raise HTTPException(status_code=400, detail="CSV must contain a 'CódigoCliente' column")
    codes: list[str] = []
    for row in reader:
        code = (row.get(field) or "").strip()
        if code:
            codes.append(code)
    if not codes:
        raise HTTPException(status_code=400, detail="No client codes found in CSV")
    return codes


def _in_memory_purchases(
    clients_csv_bytes: bytes,
    rows_total: int,
    valor_min: float,
    valor_max: float,
    inicio_str: str,
    fim_str: str,
    filiais_qtd: int,
    seed: int,
) -> tuple[bytes, bytes]:
    """Generate purchases CSV and XLSX in memory.

    Returns (csv_bytes, xlsx_bytes)
    """
    if rows_total < 1:
        raise HTTPException(status_code=400, detail="rows must be >= 1")
    if rows_total > MAX_PURCHASES:
        raise HTTPException(status_code=400, detail=f"rows exceeds limit ({MAX_PURCHASES})")
    if valor_max < valor_min:
        raise HTTPException(status_code=400, detail="valor_max must be >= valor_min")

    try:
        inicio = datetime.strptime(inicio_str, "%Y-%m-%d")
        fim = datetime.strptime(fim_str, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    if fim < inicio:
        raise HTTPException(status_code=400, detail="fim earlier than inicio")

    # Load client codes from uploaded CSV content
    codigos_clientes = _parse_client_codes_from_upload(clients_csv_bytes)

    rng = __import__("random").Random(seed)
    datas = gerar_datas(max(rows_total, len(codigos_clientes)), inicio, fim, rng)
    filiais = gerar_filiais(filiais_qtd)

    rows = montar_compras(
        codigos_clientes=codigos_clientes,
        total_linhas=rows_total,
        valor_min=valor_min,
        valor_max=valor_max,
        datas=datas,
        filiais=filiais,
        rng=rng,
    )

    header = ["CódigoCliente", "DataCompra", "Valor", "CódigoFilial"]

    # CSV
    csv_buf = io.StringIO(newline="")
    w = csv.writer(csv_buf)
    w.writerow(header)
    for cod, data, valor, filial in rows:
        w.writerow([cod, data, f"{valor:.2f}", filial])
    csv_bytes = csv_buf.getvalue().encode("utf-8")

    # XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"
    ws.append(header)
    for idx, (cod, data, valor, filial) in enumerate(rows, start=2):
        ws.append([cod, data, valor, filial])
        c = ws.cell(row=idx, column=3)
        c.number_format = "R$ #,##0.00"
    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_bytes = xlsx_buf.getvalue()

    return csv_bytes, xlsx_bytes


def _zip_bytes(files: list[tuple[str, bytes]]) -> io.BytesIO:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for name, data in files:
            z.writestr(name, data)
    buf.seek(0)
    return buf


@app.get("/", response_class=HTMLResponse)
def index(_: None = Depends(require_basic)) -> HTMLResponse:
    html_path = Path(__file__).parent / "static" / "index.html"
    content = html_path.read_text(encoding="utf-8")
    return HTMLResponse(content)


@app.post("/generate/clients")
def generate_clients(
    _: None = Depends(require_basic),
    count: int = Form(200),
    domain: str = Form("example.com"),
    seed: Optional[int] = Form(None),
):
    csv_bytes, xlsx_bytes = _in_memory_clients(count=count, domain=domain, seed=seed)
    zip_buf = _zip_bytes([
        ("clientes_ficticios.csv", csv_bytes),
        ("clientes_ficticios.xlsx", xlsx_bytes),
    ])
    headers = {"Content-Disposition": "attachment; filename=clientes.zip"}
    return StreamingResponse(zip_buf, media_type="application/zip", headers=headers)


@app.post("/generate/purchases")
async def generate_purchases(
    _: None = Depends(require_basic),
    clients_csv: UploadFile = File(...),
    rows: int = Form(5000),
    valor_min: float = Form(50.0),
    valor_max: float = Form(2000.0),
    inicio: str = Form("2024-01-01"),
    fim: str = Form("2024-12-31"),
    filiais: int = Form(5),
    seed: int = Form(42),
):
    if clients_csv.content_type not in ("text/csv", "application/vnd.ms-excel", "application/csv", None):
        # Some browsers send 'application/vnd.ms-excel' for CSV uploads
        pass
    file_bytes = await clients_csv.read()
    if len(file_bytes) > MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"CSV exceeds upload limit ({MAX_UPLOAD_MB} MB)")
    csv_bytes, xlsx_bytes = _in_memory_purchases(
        clients_csv_bytes=file_bytes,
        rows_total=rows,
        valor_min=valor_min,
        valor_max=valor_max,
        inicio_str=inicio,
        fim_str=fim,
        filiais_qtd=filiais,
        seed=seed,
    )
    zip_buf = _zip_bytes([
        ("compras.csv", csv_bytes),
        ("compras.xlsx", xlsx_bytes),
    ])
    headers = {"Content-Disposition": "attachment; filename=compras.zip"}
    return StreamingResponse(zip_buf, media_type="application/zip", headers=headers)
